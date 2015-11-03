import pandas as pd
import numpy as np
import openpyxl as xl
import re
import calendar
import psycopg2 as pg2

month_lookup = {v: k for k,v in enumerate(calendar.month_abbr)}
availablity_file_name = "C:/Users/regensbj/Documents/DTDTw/Sharepoint/SharePoint/Delivery - Digital Twin-Digital Thr/WSERInformationAndData/09.0-WSER LIMS-EV Pull and Data Sheets/LIMS-EV/Jul 15/June 30/WSER Pull 1.xlsx"
standards_file_name = "C:/Users/regensbj/Documents/DTDTw/Sharepoint/SharePoint/Delivery - Digital Twin-Digital Thr/WSERInformationAndData/09.0-WSER LIMS-EV Pull and Data Sheets/WSER Data Sheets/Jul 15/WSER Data Sheet 01Jul15.xlsx"
'''
TWSERDatabaseConnection : Python class to access and query PostgreSQL Server.

Methods:
	__init__(host_name="localhost", port_number=5432, user_name="wseruser", password=None, db_name="WSERData"):
		Initializes db connection.  Defaults to wseruser@localhost:5432.
	GetMDTable(): Returns a pandas dataframe with the available MD id and names
	GetTimeTable(): Returns a pandas dataframe with available time periods (years and quarters)
	GetMDTimeTable(): Returns a pandas dataframe unique MDs and Times.  This is the cross product of MDTable and TimeTbl
	GetAvailablityStandardsTable(): Returns the threshold by MD and year
	GetAvailablityTbl():Returns pandas dataframe for all availablity data_ending_row
	AddNewMDs: Add a new set of MDs to MDTable.  Takes TAvailablityData object and adds any MDs not current in MDTable
	AddNewTimePeriods: Add a new set of Time periods to TimeTable.  Takes TAvailablityData object and adds any time periods not current in MDTable
	LookupMDID: Gets the ID for a named MD represented as a pandas dataframe with a MD column
	LookupTimeID: Gets the ID for a named MD represented as a pandas dataframe with a MD column
	
'''
class TWSERDatabaseConnection:
    def __init__(self, host_name="localhost", port_number=5432, user_name="wseruser", password=None, db_name="WSERData"):
        self.conn = pg2.connect(database =db_name, host=host_name, port=port_number,
                                user=user_name, password=password)
    def GetMDTable(self):
        curr = self.conn.cursor()
        curr.execute("SELECT id, name FROM MDTbl;")
        results = curr.fetchall()
        df = pd.DataFrame(results, columns = ["ID", "Name"])
        curr.close()
        return df
        
    
    def GetTimeTable(self):
        curr = self.conn.cursor()
        curr.execute("SELECT id, fyyear, fyquarter FROM TimeTbl;")
        results = curr.fetchall()
        df = pd.DataFrame(results, columns = ["ID", "FYYear", "FYQuarter"])
        curr.close()
        return df        

    def GetMDTimeTable(self):
        curr = self.conn.cursor()
        curr.execute("SELECT id, md_id, time_id FROM MDTimeTbl;")
        results = curr.fetchall()
        df = pd.DataFrame(results, columns = ["ID", "MD_ID", "Time_ID"])
        curr.close()
        return df
        
    def GetAvailablityStandardsTable(self):
        curr = self.conn.cursor()
        curr.execute("SELECT MD_ID, FYear, UPNR, NMCB, NMCS, NMCM, DEPOT, AVAILABLE FROM availablitystandardstbl;")
        results = curr.fetchall()
        df = pd.DataFrame(results, columns = ["MD_ID", "FYear", "UPNR", "NMCB", "NMCS", "NMCM", "DEPOT", "AVAILABLE"])
        curr.close()
        return(df)
        
    def GetAvailablityTbl(self):
        curr = self.conn.cursor()
        curr.execute("SELECT MD_TIME_ID, UPNR, NMCB, NMCS, NMCM, DEPOT, AVAILABLE FROM AvailablityTbl;")
        results = curr.fetchall()
        df = pd.DataFrame(results, columns = ["MD_ID", "UPNR", "NMCB", "NMCS", "NMCM", "DEPOT", "AVAILABLE"])
        curr.close()
        return(df)
        
    def GetAMREPTable(self):
        curr = self.conn.cursor()
        curr.execute("SELECT id, md_id, usaf_designation, repair_org_type, completion_date FROM amreptable;")
        results = curr.fetchall()
        df = pd.DataFrame(results, columns = ["ID", "MD_ID", "USAF_DESIGNATION", "REPAIR_ORG_TYPE", "COMPLETION_DATE"])
        curr.close()
        return df
        
    
    def GetDEFECTSTable(self):
        curr = self.conn.cursor()
        curr.execute("SELECT id, md_id, usaf_designation, repair_org_type FROM defectstable;")
        results = curr.fetchall()
        df = pd.DataFrame(results, columns = ["ID", "MD_ID", "USAF_DESIGNATION", "REPAIR_ORG_TYPE"])
        curr.close()
        return df    
        
    def AddNewMDs(self, availablity_obj):
        ds_availablity = availablity_obj.GetDataframe()
        df_md = self.GetMDTable()
        values = set(ds_availablity['MD']) - set(df_md['Name'])
        curr = self.conn.cursor()
        for val in values:
            curr.execute("INSERT INTO MDTbl (Name) VALUES ('%s')"%val)
        
        self.conn.commit();
        curr.close();

    def AddNewTimePeriods(self, availablity_obj):
        time_tbl = self.GetTimeTable()       
        availablity_set = set([(availablity_obj.year, availablity_obj.quarter)])
        base_set = set([tuple(x) for x in time_tbl[['FYYear', 'FYQuarter']].values])
        new_time_points = availablity_set - base_set
        curr = self.conn.cursor()
        for (year_value, quarter_value) in new_time_points:
            curr.execute("INSERT INTO TimeTbl (FYYear,FYQuarter) VALUES (%s,%s)"%(year_value,quarter_value))
        self.conn.commit();
        curr.close();
        
    def LookupMDID(self, ds_availablity):
        df_md = self.GetMDTable()
        ds_merge = pd.merge(ds_availablity, df_md, how="left",left_on="MD", right_on="Name")
        ds_merge = ds_merge.rename(columns={'ID': 'MD_ID'})
        return ds_merge
        
    def LookupTimeID(self, ds_availablity):
        df_time = self.GetTimeTable()
        ds_merge = pd.merge(ds_availablity, df_time, how="left",left_on=["Year", "Quarter"], right_on=["FYYear", "FYQuarter"])
        ds_merge = ds_merge.rename(columns={'ID': 'Time_ID'})
        return ds_merge

        
    def AddNewMDTimes(self, availablity_obj):
        self.AddNewMDs(availablity_obj)
        self.AddNewTimePeriods(availablity_obj)
        ds_availablity = availablity_obj.GetDataframe()
        ds_merge = self.LookupMDID(ds_availablity)
        ds_merge = self.LookupTimeID(ds_merge)
        
        ds_merge = ds_merge[['MD_ID', 'Time_ID']]
        ds_merge = ds_merge.drop_duplicates()
        md_time_tbl = self.GetMDTimeTable()
        
        availablity_set = set([tuple(x) for x in ds_merge.values])
        base_set = set([tuple(x) for x in md_time_tbl[["MD_ID", "Time_ID"]].values])
        new_time_points = availablity_set - base_set
        curr = self.conn.cursor()
        for (md_value, time_value) in new_time_points:
            curr.execute("INSERT INTO MDTimeTbl (MD_ID, Time_ID) VALUES (%s,%s)"%(md_value,time_value))
        self.conn.commit();
        curr.close();
        
    def AppendMDSTimeIDToDataFrame(self, ds_availablity):
        ds_merge = self.LookupMDID(ds_availablity)
        ds_merge = self.LookupTimeID(ds_merge)
        md_time_tbl = self.GetMDTimeTable()
        ds_merge = pd.merge(ds_merge, md_time_tbl, how="left",on=["MD_ID", "Time_ID"])
        ds_merge = ds_merge.rename(columns={'ID': 'MD_Time_ID'})   
        ds_merge = ds_merge[["MD_Time_ID",  "AVAILABLE", "DEPOT", "NMCB", "NMCM", "NMCS", "UPNR"]]
        return(ds_merge)
        
    def AppendMDSToDataFrameAmrep(self, data_frame):
        ds_merge = self.LookupMDID(data_frame)
        ds_merge = ds_merge.rename(columns={'ID': 'MD_ID'})   
        ds_merge = ds_merge[["MD_ID",  "REPAIR_ORG_TYPE", "USAF_DESIGNATION", "COMPLETION_DATE"]]
        return(ds_merge)
        
    def AppendMDSToDataFrameDefects(self, data_frame):
        ds_merge = self.LookupMDID(data_frame)
        ds_merge = ds_merge.rename(columns={'ID': 'MD_ID'})   
        ds_merge = ds_merge[["MD_ID",  "REPAIR_ORG_TYPE", "USAF_DESIGNATION", "AC_DATE"]]
        return(ds_merge)
        
    def AppendAvailablityStandards(self, standards_obj):
        standards_df = standards_obj.ToDataFrame()
        md_df = self.GetMDTable()
        merged_data = pd.merge(standards_df, md_df, how='left', left_on='MD', right_on='Name')
        is_valid = np.isfinite(merged_data['ID'])
        missing_mds = merged_data[list(~is_valid)]['MD'].unique()
        print("No MD ID found for the following standards: %s"%missing_mds) 
        valid_standards = merged_data[list(is_valid)]
        valid_standards = valid_standards.rename(columns={'ID': 'MD_ID'})
        df_existing_standards = self.GetAvailablityStandardsTable()
        new_entries = list(pd.merge(valid_standards, df_existing_standards, 'left', left_on = ['MD_ID', 'Year'], right_on = ['MD_ID', 'FYear'])['FYear'].isnull())
        entry_counter = 0
        curr = self.conn.cursor()
        for ii, row in valid_standards.iterrows():
            if new_entries[entry_counter]:
                if (row.notnull().all()):
                    curr.execute("INSERT INTO availablitystandardstbl (MD_ID, FYear, UPNR, NMCB, NMCS, NMCM, DEPOT, AVAILABLE) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"%(row['MD_ID'], 
                                 row['Year'], row['UPNR'], row['NMCB'], row['NMCS'], row['NMCM'], row['Depot'], row['Available']))
            else:
                print("Duplicate Entry Found %s, %s:FY%i"%(row['MD_ID'], row['Name'], row['Year']))
            entry_counter += 1
        self.conn.commit()
        curr.close();
                         

    def AddNewAvailablityData(self, availablity_obj):
        self.AddNewMDTimes(availablity_obj)
        ds_availablity = availablity_obj.GetDataframe()
        tmp_x = self.AppendMDSTimeIDToDataFrame(ds_availablity)
        curr = self.conn.cursor()
        df_stored_values = self.GetAvailablityTbl()
        new_mds_time_ids = set(tmp_x['MD_Time_ID']) - set(df_stored_values['MD_ID'])
        check = tmp_x['MD_Time_ID'].isin(new_mds_time_ids)
        new_mds_time_df = tmp_x[check]

        if not all(check):
            idx = tmp_x[check==False].index
            duplicate_indices = idx.get_values()
            print("Duplicates entries found for the following rows, skipping (%s)"%duplicate_indices)        
            
            
        if any(check):
            for ii, row in new_mds_time_df.iterrows():
                curr.execute("INSERT INTO AvailablityTbl (MD_TIME_ID, UPNR, NMCB, NMCS, NMCM, DEPOT, AVAILABLE) VALUES (%s, %s, %s, %s, %s, %s, %s)"%(row['MD_Time_ID'], 
                             row['UPNR'], row['NMCB'], row['NMCS'], row['NMCM'], row['DEPOT'], row['AVAILABLE']))
                         

        self.conn.commit()
        curr.close();
        
    def AddNewAmrepData(self, amrep_table):
        self.AddNewMDs(amrep_table)
        amrep_table = amrep_table.GetDataframe()
        tmp_x = self.AppendMDSToDataFrameAmrep(amrep_table)
        curr = self.conn.cursor()
        for ii, row in tmp_x.iterrows():
            query = "INSERT INTO amreptable (MD_ID, USAF_DESIGNATION, REPAIR_ORG_TYPE, COMPLETION_DATE) VALUES (%s, %s, %s, %s)"
            values = (row['MD_ID'], row['USAF_DESIGNATION'], row['REPAIR_ORG_TYPE'], row['COMPLETION_DATE'])
            curr.execute(query, values)            
        self.conn.commit()
        curr.close();
        
    def AddNewDefectsData(self, defects_table):
        self.AddNewMDs(defects_table)
        defects_table = defects_table.GetDataframe()
        tmp_x = self.AppendMDSToDataFrameDefects(defects_table)
        curr = self.conn.cursor()
        for ii, row in tmp_x.iterrows():
            query = "INSERT INTO defectstable (MD_ID, USAF_DESIGNATION, REPAIR_ORG_TYPE, AC_DATE) VALUES (%s, %s, %s, %s)"
            values = (row['MD_ID'], row['USAF_DESIGNATION'], row['REPAIR_ORG_TYPE'], row['AC_DATE'])
            curr.execute(query, values)    
        self.conn.commit()
        curr.close();

    def Close(self):
        self.conn.close()
        

class TAvailablityData:
    def __init__(self, file_name):
        tmp_dictionary = {"DataFrameLabel": ["MD", "UPNR", "NMCB", "NMCS", "NMCM", "DEPOT", "AVAILABLE"],
        "ColumnLabel": ["MD", "UPNR(%)","NMCB(NA)(%)", "NMCS(NA)(%)", "NMCM(NA)(%)", "Depot(%)", "Available(%)"], 
        "ColumnIndex":[None]*7,
        "Values":[None]*7}
        self.label_dictionary = pd.DataFrame(tmp_dictionary)      
        self.label_dictionary = self.label_dictionary.set_index("DataFrameLabel")
        self.source_file = file_name
        self.year = None
        self.quarter = None        
        self.data_frame = None        
        self.ReadData()
        
    def ReadData(self):
        wb = xl.load_workbook(self.source_file)

        if not "Sheet1" in wb.get_sheet_names():
            raise ValueError("(%s) missing sheet name"%self.source_file)
        
        ws = wb.get_sheet_by_name("Sheet1")
        self.PopulateHeader(ws)        
        self.PopulateDateFrame(ws)
        
    def PopulateHeader(self, ws):
        row_array = ws.rows
        for rx in row_array:
            tmp_string = rx[0].value
            if not tmp_string==None:
                starting_index = tmp_string.find("Date Range:")
                check = starting_index >= 0
                if check:
                    search_results = re.findall("Date Range: By Month, From (.*) (.*) to (.*) (.*)", tmp_string[starting_index:])
                    if len(search_results)==1:
                        month_value = month_lookup[search_results[0][2]]
                        self.year = int(search_results[0][3])
                
                        if (month_value > 10):
                            self.year += 1
                            month_value -= 12            
                        self.quarter = int(4*month_value/12)+1                
                    break;
                
    
    def PopulateDateFrame(self, ws):
        row_array = ws.rows
        starting_row_index = self.GetStartingDataRow(row_array)
        if starting_row_index < 0:
            raise ValueError("(%s) does not contain the expect column headings"%self.source_file)
        
        num_rows = len(row_array)   
        num_data_rows = num_rows - starting_row_index  
        for idx in self.label_dictionary.index:
            self.label_dictionary['Values'][idx] = [None] * num_data_rows
        
        for row_index in range(starting_row_index+1, num_rows):
            offset_row = row_index - starting_row_index - 1
            row = row_array[row_index]
            column_indices = self.label_dictionary['ColumnIndex']
            for ii, col_idx in enumerate(column_indices):
                if type(col_idx) is int:
                    tmp_value = row[col_idx].value
                    self.label_dictionary['Values'][ii][offset_row] = tmp_value
       
    def GetStartingDataRow(self, row_array):
        starting_row = -1
        
        for row_index, rx in enumerate(row_array):
            starting_row_found = False                 
            for col_index, cx in enumerate(rx):
                tmp_string = cx.value
                if type(tmp_string) is str:
                    col_name = tmp_string.replace(" ", "")
                    starting_row_found |= col_name == "MD"
                    check = self.label_dictionary['ColumnLabel']==col_name
                    if any(check):
                        self.label_dictionary['ColumnIndex'][check] = col_index
                        starting_row_found = True

            if starting_row_found:
                starting_row = row_index
                break;
                
        return starting_row
            
    def GetDataframe(self):
        tmp_dict = {}
        valid_entry_index = [idx for idx, value in enumerate(self.label_dictionary['Values']['MD']) if value is not None]
        for idx in self.label_dictionary.index:            
            tmp_dict[idx] = list(self.label_dictionary['Values'][idx][i] for i in valid_entry_index )
        
        df = pd.DataFrame(tmp_dict)
        df['Year'] = self.year;
        df['Quarter'] = self.quarter;
        return df;
        


class TAvailablityStandardsData:
    def __init__(self, file_name):
        tmp_dictionary = {"DataFrameLabel": ["MD", "UPNR", "NMCB", "NMCS", "NMCM", "DEPOT", "AVAILABLE"],
        "ColumnLabel": ["MD", "UPNR(%)","NMCB(NA)(%)", "NMCS(NA)(%)", "NMCM(NA)(%)", "Depot(%)", "Available(%)"], 
        "ColumnIndex":[None]*7,
        "Values":[None]*7}
        self.label_dictionary = pd.DataFrame(tmp_dictionary)      
        self.label_dictionary = self.label_dictionary.set_index("DataFrameLabel")
        self.source_file = file_name
        self.year = None
        self.quarter = None    
        self.data_frame = None        
        self.ReadData()
        
    def ReadData(self):
        wb = xl.load_workbook(self.source_file)
        if not "Sheet1" in wb.get_sheet_names():
            raise ValueError("(%s) missing sheet name"%self.source_file)
        
        ws = wb.get_sheet_by_name("Sheet1")
        self.PopulateHeader(ws)        
        self.PopulateDateFrame(ws)
        
    def PopulateHeader(self, ws):
        row_array = ws.rows
        for rx in row_array:
            tmp_string = rx[0].value
            if not tmp_string==None:
                starting_index = tmp_string.find("Date Range:")
                check = starting_index >= 0
                if check:
                    search_results = re.findall("Date Range: By Month, From (.*) (.*) to (.*) (.*)", tmp_string[starting_index:])
                    if len(search_results)==1:
                        month_value = month_lookup[search_results[0][2]]
                        self.year = int(search_results[0][3])
                
                        if (month_value > 10):
                            self.year += 1
                            month_value -= 12            
                        self.quarter = int(4*month_value/12)+1                
                    break;
                
    
    def PopulateDateFrame(self, ws):
        row_array = ws.rows
        starting_row_index = self.GetStartingDataRow(row_array)
        if starting_row_index < 0:
            raise ValueError("(%s) does not contain the expect column headings"%self.source_file)
        
        num_rows = len(row_array)   
        num_data_rows = num_rows - starting_row_index  
        for idx in self.label_dictionary.index:
            self.label_dictionary['Values'][idx] = [None] * num_data_rows
        
        for row_index in range(starting_row_index+1, num_rows):
            offset_row = row_index - starting_row_index - 1
            row = row_array[row_index]
            column_indices = self.label_dictionary['ColumnIndex']
            for ii, col_idx in enumerate(column_indices):
                if type(col_idx) is int:
                    if (row[col_idx].value):
                        tmp_value = row[col_idx].value
                        self.label_dictionary['Values'][ii][offset_row] = tmp_value
       
    def GetStartingDataRow(self, row_array):
        starting_row = -1
        
        for row_index, rx in enumerate(row_array):
            starting_row_found = False                 
            for col_index, cx in enumerate(rx):
                tmp_string = cx.value
                if type(tmp_string) is str:
                    col_name = tmp_string.replace(" ", "")
                    starting_row_found |= col_name == "MD"
                    check = self.label_dictionary['ColumnLabel']==col_name
                    if any(check):
                        self.label_dictionary['ColumnIndex'][check] = col_index

            if starting_row_found:
                starting_row = row_index
                break;
                
        return starting_row
            
    def GetDataframe(self):
        tmp_dict = {}
        valid_entry_index = [idx for idx, value in enumerate(self.label_dictionary['Values']['MD']) if value is not None]
        for idx in self.label_dictionary.index:            
            tmp_dict[idx] = list(self.label_dictionary['Values'][idx][i] for i in valid_entry_index )
        
        df = pd.DataFrame(tmp_dict)
        df['Year'] = self.year;
        df['Quarter'] = self.quarter;
        return df;    
 
   
    
#def main():  
#    df_avail = ImportAvailablityData(availablity_file_name) 
#    db_connection = TWSERDatabaseConnection(password="wseruser");
#    db_connection.AddNewAvailablityData(df_avail)
#    db_connection.Close()
#if __name__ == "__main__":
#    main()

def ImportAvailablityData(file_name):
    # Add check if file exists

    
    #if not "Sheet1" in sheet_names:
        # Add error
        

    header_size = 6
    num_data_rows = num_rows - header_size - 2# Skip last two rows
    num_data_cols = 6
    
    time_period_row_index = 2
    time_period_row = row_array[time_period_row_index]
    time_period_text = time_period_row[0].value
    search_results = re.findall("Date Range: By Month, From (.*) (.*) to (.*) (.*)", time_period_text)
    year_value = None
    quarter_value = None
    
    if len(search_results)==1:
        month_value = month_lookup[search_results[0][2]]
        year_value = int(search_results[0][3])

        if (month_value > 10):
            year_value += 1
            month_value -= 12

        quarter_value = int(4*month_value/12)+1
        
    label_array = [None] * num_data_rows
    value_matrix = np.zeros((num_data_rows, num_data_cols))
    data_starting_row = header_size
    data_ending_row = header_size + num_data_rows
    for row_index in range(data_starting_row, data_ending_row):
        row_offset = row_index - data_starting_row
        row = row_array[row_index]
        label_array[row_offset] = row[0].value
        for col_index in range(1, num_data_cols+1):
            col_offset = col_index - 1
            value_matrix[row_offset, col_offset] = row[col_index].value
        
    ds = {"MD": label_array, "UPNR": value_matrix[:, 0], 
    "NMCB": value_matrix[:, 1], "NMCS": value_matrix[:, 2], 
    "NMCM": value_matrix[:, 3], "Depot": value_matrix[:, 4], 
    "Available": value_matrix[:, 5], 
    "Year": [year_value]*num_data_rows, 
    "Quarter": [quarter_value] * num_data_rows}
    ds = pd.DataFrame(ds)
    return(ds)
    
class AmrepTableData:
        def __init__(self, file_name):
            tmp_dictionary = {"DataFrameLabel": ["MD", "USAF_DESIGNATION", "REPAIR_ORG_TYPE", "COMPLETION_DATE"],
            "ColumnLabel": ["MD", "USAF", "REPAIRORGTYPE", "COMPL"], 
            "ColumnIndex":[None]*4,
            "Values":[None]*4}
            self.label_dictionary = pd.DataFrame(tmp_dictionary)      
            self.label_dictionary = self.label_dictionary.set_index("DataFrameLabel")
            self.source_file = file_name      
            self.data_frame = None        
            self.ReadData()
            
        def ReadData(self):
            wb = xl.load_workbook(self.source_file, data_only = True)
    
            if not "AMREP" in wb.get_sheet_names():
                raise ValueError("(%s) missing sheet name"%self.source_file)
            
            ws = wb.get_sheet_by_name("AMREP")      
            self.PopulateDateFrame(ws)
        
        def PopulateDateFrame(self, ws):
            row_array = ws.rows
            starting_row_index = self.GetStartingDataRow(row_array)
            if starting_row_index < 0:
                raise ValueError("(%s) does not contain the expect column headings"%self.source_file)
            
            num_rows = len(row_array)   
            num_data_rows = num_rows - starting_row_index  
            for idx in self.label_dictionary.index:
                self.label_dictionary['Values'][idx] = [None] * num_data_rows
            for row_index in range(starting_row_index+1, num_rows):
                offset_row = row_index - starting_row_index - 1
                row = row_array[row_index]
                column_indices = self.label_dictionary['ColumnIndex']
                for ii, col_idx in enumerate(column_indices):
                    if type(col_idx) is int:
                        if (row[col_idx].value):
                            tmp_value = row[col_idx].value
                            self.label_dictionary['Values'][ii][offset_row] = tmp_value
                        
           
        def GetStartingDataRow(self, row_array):
            starting_row = -1
            
            for row_index, rx in enumerate(row_array):
                starting_row_found = False                 
                for col_index, cx in enumerate(rx):
                    tmp_string = cx.value
                    if type(tmp_string) is str:
                        col_name = tmp_string.replace(" ", "")
                        starting_row_found |= col_name == "MD"
                        check = self.label_dictionary['ColumnLabel']==col_name
                        if any(check):
                            self.label_dictionary['ColumnIndex'][check] = col_index
                            starting_row_found = True
    
                if starting_row_found:
                    starting_row = row_index
                    break;
                    
            return starting_row
                
        def GetDataframe(self):
            tmp_dict = {}
            valid_entry_index = [idx for idx, value in enumerate(self.label_dictionary['Values']['MD']) if value is not None]
            for idx in self.label_dictionary.index:            
                tmp_dict[idx] = list(self.label_dictionary['Values'][idx][i] for i in valid_entry_index )
            
            df = pd.DataFrame(tmp_dict)
            return df;
            
class DefectsTableData:
        def __init__(self, file_name):
            tmp_dictionary = {"DataFrameLabel": ["MD", "USAF_DESIGNATION", "REPAIR_ORG_TYPE", "AC_DATE"],
            "ColumnLabel": ["MD", "AirForce", "ORG/CONT", "A/CCompleted"], 
            "ColumnIndex":[None]*4,
            "Values":[None]*4}
            self.label_dictionary = pd.DataFrame(tmp_dictionary)      
            self.label_dictionary = self.label_dictionary.set_index("DataFrameLabel")
            self.source_file = file_name      
            self.data_frame = None        
            self.ReadData()
            
        def ReadData(self):
            wb = xl.load_workbook(self.source_file, data_only = True)

            if not "Defects" in wb.get_sheet_names():
                raise ValueError("(%s) missing sheet name"%self.source_file)
            
            ws = wb.get_sheet_by_name("Defects")       
            self.PopulateDateFrame(ws)
        
        def PopulateDateFrame(self, ws):
            row_array = ws.rows
            starting_row_index = self.GetStartingDataRow(row_array)
            if starting_row_index < 0:
                raise ValueError("(%s) does not contain the expect column headings"%self.source_file)
            
            num_rows = len(row_array)   
            num_data_rows = num_rows - starting_row_index  
            for idx in self.label_dictionary.index:
                self.label_dictionary['Values'][idx] = [None] * num_data_rows
            for row_index in range(starting_row_index+1, num_rows):
                offset_row = row_index - starting_row_index - 1
                row = row_array[row_index]
                column_indices = self.label_dictionary['ColumnIndex']
                for ii, col_idx in enumerate(column_indices):
                    if type(col_idx) is int:   
                        if (row[col_idx].value):
                            tmp_value = row[col_idx].value
                            if (tmp_value == 'C'):
                                tmp_value = "CON"
                            if (tmp_value == 'O'):
                                tmp_value = "ORG"   
                            self.label_dictionary['Values'][ii][offset_row] = tmp_value
            
        def GetStartingDataRow(self, row_array):
            starting_row = -1
            
            for row_index, rx in enumerate(row_array):
                starting_row_found = False                 
                for col_index, cx in enumerate(rx):
                    tmp_string = cx.value
                    if type(tmp_string) is str:
                        col_name = tmp_string.replace(" ", "")
                        starting_row_found |= col_name == "MD"
                        check = self.label_dictionary['ColumnLabel']==col_name
                        if any(check):
                            self.label_dictionary['ColumnIndex'][check] = col_index
                            starting_row_found = True
    
                if starting_row_found:
                    starting_row = row_index
                    break;
                    
            return starting_row
                
        def GetDataframe(self):
            tmp_dict = {}
            valid_entry_index = [idx for idx, value in enumerate(self.label_dictionary['Values']['MD']) if value is not None]
            for idx in self.label_dictionary.index:            
                tmp_dict[idx] = list(self.label_dictionary['Values'][idx][i] for i in valid_entry_index )
            
            df = pd.DataFrame(tmp_dict)
            return df;
